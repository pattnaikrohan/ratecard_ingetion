import openpyxl
import re
from pathlib import Path
from typing import List
from app.services.parsers.base_parser import BaseParser
from app.models.canonical import CanonicalRateSheet, RateRow, ChargeItem, JobSummary

class ONEPlugin(BaseParser):
    def can_parse(self, file_path: Path, filename: str) -> bool:
        fn = filename.lower()
        return "one " in fn or "oney" in fn or "ocean network" in fn or "one-line" in fn

    def parse(self, file_path: Path, job_id: str) -> CanonicalRateSheet:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        rates: List[RateRow] = []
        row_counter = 1
        validity_start = ""
        validity_end = ""
        contract_number = ""

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 6:
                continue

            # Read top rows for metadata
            for top_r in range(1, 6):
                row_str = " ".join(str(ws.cell(top_r, c).value or "") for c in range(1, 15))
                if "MRG" in row_str or "Contract" in row_str:
                    m = re.search(r'([A-Za-z0-9\-_]{4,})', row_str)
                    if m and not contract_number:
                        contract_number = m.group(1)
                
                # Check for Validity or Effective/Expiry
                if re.search(r'(valid|effect|expir)', row_str, re.IGNORECASE):
                    dates = re.findall(r'(\d{2,4}[-/]\d{1,2}[-/]\d{1,4}|\d{1,2}-[a-zA-Z]{3}-\d{2,4})', row_str)
                    if len(dates) >= 1 and not validity_start:
                        validity_start = dates[0]
                    if len(dates) >= 2 and not validity_end:
                        validity_end = dates[1]

            # Read headers from row 3
            headers = [str(ws.cell(3, c).value or "").strip() for c in range(1, ws.max_column + 1)]
            
            por_idx = self._find_col(headers, "POR")
            pod_idx = self._find_col(headers, "POD")
            svc_idx = self._find_col(headers, "Service")

            # Map load type columns
            equipment_cols = []
            for c_idx, h in enumerate(headers):
                if h in ["20'", "40'", "40'HC", "20'RF", "40'RF", "20'RAD", "40'RAD"]:
                    equipment_cols.append((c_idx + 1, h))

            for r in range(6, ws.max_row + 1):
                por_val = str(ws.cell(r, por_idx).value or "").strip()
                pod_val = str(ws.cell(r, pod_idx).value or "").strip()
                svc_val = str(ws.cell(r, svc_idx).value or "").strip()

                if not por_val or not pod_val:
                    continue

                # Expand combined PODs (e.g. "Brisbane / Melbourne / Adelaide" -> ["Brisbane", "Melbourne", "Adelaide"])
                pods = [p.strip() for p in pod_val.split("/") if p.strip()]

                for pod_item in pods:
                    for col_num, eq_header in equipment_cols:
                        amt_val = ws.cell(r, col_num).value
                        if amt_val is not None and str(amt_val).strip() != "":
                            try:
                                amt = float(amt_val)
                            except ValueError:
                                continue

                            load_type = self._map_equipment(eq_header)
                            
                            r_row = RateRow(
                                row_index=row_counter,
                                carrier_scac="ONEY",
                                origin_raw=por_val,
                                origin_locode=por_val,
                                destination_raw=pod_item,
                                destination_locode=pod_item,
                                service_type=svc_val,
                                cargo_type="FAK",
                                load_type=load_type,
                                commodity="FAK",
                                ofr_amount=amt,
                                ofr_currency="USD",
                                charges=[ChargeItem(charge_code="OFR", charge_name="Base Ocean Freight", amount=amt, currency="USD", basis="per equipment")],
                                validity_start=validity_start,
                                validity_end=validity_end,
                                contract_number=contract_number,
                                inclusions="",
                                subject_to="",
                                remarks=""
                            )
                            rates.append(r_row)
                            row_counter += 1

        summary = JobSummary(total_rows=len(rates), carriers_found=["ONEY"])
        return CanonicalRateSheet(
            job_id=job_id,
            file_name=file_path.name,
            carrier_code="ONEY",
            contract_number=contract_number,
            validity_start=validity_start,
            validity_end=validity_end,
            rates=rates,
            summary=summary
        )

    def _find_col(self, headers: List[str], target: str) -> int:
        for idx, h in enumerate(headers):
            if target.lower() == h.lower():
                return idx + 1
        return 1

    def _map_equipment(self, header: str) -> str:
        mapping = {
            "20'": "20GP",
            "40'": "40GP",
            "40'HC": "40HC",
            "20'RF": "20RF",
            "40'RF": "40RF",
            "20'RAD": "20GP",
            "40'RAD": "40HC"
        }
        return mapping.get(header, "20GP")

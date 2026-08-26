"""
Generic Excel Rate Card Parser
Intelligently auto-detects column headers and extracts rate data from any carrier's Excel format.
Supports multi-sheet workbooks, multi-tab surcharge merging (e.g. Summary+Standard charges, Rate+Surcharge),
multi-port cell splitting (newline & semicolon separated origins/destinations), and junk/lookup sheet filtering.
"""
import openpyxl
import re
import datetime
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any
from app.services.parsers.base_parser import BaseParser
from app.models.canonical import CanonicalRateSheet, RateRow, ChargeItem, JobSummary
from app.services.ai_column_mapper import AIColumnMapper
from app.utils_contracts import is_valid_contract_number

# ── Sheet Blacklist ─────────────────────────────────────────────────────────
IGNORED_SHEET_NAMES = {
    "search", "default", "default1", "default2", "abbreviation", "abbreviations",
    "lookup", "dropdown", "dropdowns", "title page", "titlepage", "cover",
    "instruction", "instructions", "disclaimer", "general disclaimer",
    "template guide", "read me", "readme", "notes", "legend", "formula", "formulas"
}

SURCHARGE_SHEET_KEYWORDS = {
    "standard charges", "standard charge", "surcharge", "surcharges",
    "local charges", "additional charges", "ancillary charges", "ancillaries"
}

ORIGIN_SYNONYMS = {
    "por", "pol", "origin", "origins", "from", "port of loading", "load port",
    "loading port", "port of origin", "cntr origin", "origin port",
    "departure", "dep", "dep port", "ex", "source", "receipt", "place of receipt",
    "origin city", "origin inland cfs", "consol cfs",
}

DESTINATION_SYNONYMS = {
    "pod", "destination", "destinations", "dest", "to", "port of discharge", "discharge port",
    "disch port", "dest port", "delivery", "del", "del port", "arrival",
    "destination port", "arr", "place of delivery", "destination city",
    "destination inland cfs", "deconsol cfs",
}

ORIGIN_CODE_SYNONYMS = {"port code", "pol code", "origin code", "origin locode", "origin un code"}
DEST_CODE_SYNONYMS = {"destination code", "dest code", "pod code", "destination locode", "destination un code", "port of discharge code"}
CARRIER_SYNONYMS = {"carrier", "scac", "line", "shipping line", "carrier scac", "carrier code", "operator"}
CURRENCY_SYNONYMS = {"currency", "curr", "cur", "ccy"}
SERVICE_SYNONYMS = {"service", "svc", "service type", "mode", "routing"}
VALIDITY_START_SYNONYMS = {"effective", "valid from", "validity from", "start", "eff date", "effective date", "from date", "valid date", "rate effective date", "valid fm", "validity start", "eff_date", "eff"}
VALIDITY_END_SYNONYMS = {"expiry", "valid to", "validity to", "end", "exp date", "expiry date", "to date", "until", "validity", "rate expiry date", "validity end", "exp_date", "exp"}
CONTRACT_SYNONYMS = {"contract", "service contract", "sc no", "contract no", "contract number", "sc number", "agreement"}
COMMODITY_SYNONYMS = {"commodity", "comm", "cargo", "goods", "commodity type", "cargo nature"}
TRANSIT_TIME_SYNONYMS = {"transit time", "transit", "tt", "t/t", "days", "estimated transit time"}
CHARGE_CODE_SYNONYMS = {"charge code", "charge", "chg code"}
CHARGE_NAME_SYNONYMS = {"charge name", "chg name", "description", "charge description"}
AMOUNT_SYNONYMS = {"amount", "rate", "charge amount", "amt"}
RATE_BASIS_SYNONYMS = {"rate basis", "basis", "per", "uom", "unit"}

CONTAINER_PATTERNS: List[Tuple[str, List[str]]] = [
    ("20DG", ["20'dg", "20dg", "20' hazardous", "20 hazardous"]),
    ("40DG", ["40'dg", "40dg", "40' hazardous", "40 hazardous"]),
    ("45DG", ["45'dg", "45dg", "45' hazardous", "45 hazardous"]),
    ("20OT", ["20'ot", "20ot", "20'ot/fr", "20ot/fr"]),
    ("40OT", ["40'ot", "40ot", "40'ot/fr", "40ot/fr"]),
    ("20FR", ["20'fr", "20fr", "20 flat rack"]),
    ("40FR", ["40'fr", "40fr", "40 flat rack"]),
    ("40HC", ["40'hc", "40hc", "40'hq", "40hq", "40hi", "h40", "40hdry", "40hdr", "rate 40h", "rate 40hc", "rate 40hq", "40h", "40hq"]),
    ("45HC", ["45'hc", "45hc", "45'hq", "45hq", "rate 45", "rate 45h", "rate 45hc"]),
    ("45GP", ["45'gp", "45gp", "45'st", "45'dv", "45dv", "45ft", "45dr"]),
    ("20RF", ["20'rf", "20rf", "20reefer", "r20"]),
    ("40RF", ["40'rf", "40rf", "40reefer", "r40"]),
    ("20NOR", ["20nor", "20'nor"]),
    ("40NOR", ["40nor", "40'nor"]),
    ("20GP", ["20'st", "20'gp", "20gp", "20'dv", "20dv", "20ft", "20dr", "d20", "20dry", "20st", "20'", "rate 20", "rate 20st", "rate 20gp"]),
    ("40GP", ["40'st", "40'gp", "40gp", "40'dv", "40dv", "40ft", "40dr", "d40", "40dry", "40st", "40'", "rate 40", "rate 40st", "rate 40gp"]),
    ("LCL",  ["lcl", "cbm", "w/m", "per cbm", "per w/m"]),
]

LCL_RATE_SYNONYMS = {
    "rate per cbm", "rate per tonne", "rate per w/m", "ocean freight rate",
    "ocean freight", "ofr", "minimum charge", "rate", 
    "per cbm", "per ton", "min", "w/m", "rt",
}

def _normalize(s: str) -> str:
    s_clean = re.sub(r'[*:#\(\)]', ' ', s)
    return re.sub(r'\s+', ' ', s_clean.strip()).lower()

def _match_synonym(header: str, synonyms: set) -> bool:
    h = _normalize(header)
    if h in synonyms:
        return True
    for syn in synonyms:
        if len(syn) > 4 and syn in h:
            return True
        elif 3 <= len(syn) <= 4:
            if re.search(r'(?:^|[\s\-/,])' + re.escape(syn) + r'(?:$|[\s\-/,])', h):
                return True
    return False

def _match_container_type(header: str) -> Optional[str]:
    h = _normalize(header)
    for std_type, patterns in CONTAINER_PATTERNS:
        for pat in patterns:
            if pat == h:
                return std_type
            if re.search(r'(?:^|[\s\-_/])' + re.escape(pat) + r'(?:$|[\s\-_/])', h):
                return std_type
    return None

def _match_lcl_rate(header: str) -> bool:
    h = _normalize(header)
    if h in LCL_RATE_SYNONYMS:
        return True
    for syn in LCL_RATE_SYNONYMS:
        if len(syn) > 4 and syn in h:
            return True
    return False

def _extract_amount(val) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.upper() in ('N/A', 'NA', '-', 'TBA', 'TBC', 'POA', 'INCL', 'INCLUSIVE', 'INCLUDED', 'NONE', 'NULL'):
        return None
    s = s.replace('$', '').replace('€', '').replace('£', '').replace('¥', '')
    s = s.replace(',', '').replace(' ', '')
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    try:
        amt = float(s)
        return amt if amt > 0 else None
    except ValueError:
        digits = re.findall(r'\d+\.?\d*', s)
        if digits:
            try:
                amt = float(digits[0])
                return amt if amt > 0 else None
            except ValueError:
                return None
    return None

def _clean_date(date_val) -> str:
    if not date_val:
        return ""
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        return date_val.strftime("%Y-%m-%d")
    s = str(date_val).strip()
    s = re.sub(r'\s+\d{2}:\d{2}:\d{2}.*$', '', s)
    for fmt in ["%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y",
                "%d %b %Y", "%d %B %Y", "%Y/%m/%d", "%b %d, %Y", "%B %d, %Y",
                "%Y-%m-%d %H:%M:%S"]:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s

def _is_junk_rate_row(origin: str, dest: str, ofr: float) -> bool:
    o_low = origin.lower().strip()
    d_low = dest.lower().strip()
    
    header_keywords = [
        "port of loading", "port of discharge", "place of receipt", "place of delivery",
        "load port", "discharge port", "origin group", "destination group",
        "origin:", "destination:", "origin group:", "destination group:",
        "click for details", "standard", "disclaimer", "subject to", "comment",
        "cntr type", "cargo nature", "tariff involved", "alternative name",
        "quote line", "charge description", "container charges", "rate offer"
    ]
    if any(k in o_low for k in header_keywords) or any(k in d_low for k in header_keywords):
        return True
    if o_low in ("load port", "discharge port", "por", "pod", "pol"):
        return True
    if re.match(r'^(?:20|40|45)(?:gp|hc|st|dry|ft)?$', o_low) or re.match(r'^(?:20|40|45)(?:gp|hc|st|dry|ft)?$', d_low):
        return True
    return False


class GenericExcelPlugin(BaseParser):
    def can_parse(self, file_path: Path, filename: str) -> bool:
        fn = filename.lower()
        return fn.endswith('.xlsx') or fn.endswith('.xls') or fn.endswith('.xlsm')

    def parse(self, file_path: Path, job_id: str) -> CanonicalRateSheet:
        print(f"[GenericExcel] Parsing: {file_path.name} (Job: {job_id})")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_rates: List[RateRow] = []
        row_counter = 1
        detected_carrier = ""
        detected_validity_start = ""
        detected_validity_end = ""
        detected_contract = ""
        detected_currency = "USD"

        # Carrier detection from filename
        fn = file_path.name
        carrier_from_filename = self._detect_carrier_from_filename(fn)
        validity_from_filename = self._detect_validity_from_filename(fn)
        if validity_from_filename:
            detected_validity_start, detected_validity_end = validity_from_filename

        # Filter out blacklisted sheets
        active_sheet_names = [
            s for s in wb.sheetnames
            if _normalize(s) not in IGNORED_SHEET_NAMES
        ]

        if not active_sheet_names:
            active_sheet_names = wb.sheetnames

        # Surcharge Sheets detection
        surcharge_sheets = [s for s in active_sheet_names if any(k in _normalize(s) for k in SURCHARGE_SHEET_KEYWORDS)]
        rate_sheets = [s for s in active_sheet_names if s not in surcharge_sheets]
        if not rate_sheets:
            rate_sheets = active_sheet_names

        surcharges_cache = self._extract_surcharge_sheet_data(wb, surcharge_sheets)
        print(f"[GenericExcel] Active rate sheets: {rate_sheets}, Surcharge sheets: {surcharge_sheets} (Cached: {len(surcharges_cache)} routes)")

        for sheet_name in rate_sheets:
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                continue

            max_col = min(ws.max_column or 1, 50)
            max_row = min(ws.max_row or 1, 5000)

            # Metadata scan from top rows
            for r in range(1, min(max_row + 1, 15)):
                row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, max_col + 1))
                if not detected_validity_start and re.search(r'validity|effective|valid from', row_text, re.IGNORECASE):
                    text_clean = re.sub(r'\s+', ' ', row_text).strip()
                    m = re.search(r'(?:validity|effective)?[:\s]*(\d{1,2}\s+[a-zA-Z]{3,9}(?:\s+\d{4})?|\d{1,2}[\-/][a-zA-Z]{3,9}[\-/]\d{2,4}|\d{4}[\-/]\d{1,2}[\-/]\d{1,2})\s*(?:to|\-|\~)\s*(\d{1,2}\s+[a-zA-Z]{3,9}(?:\s+\d{4})?|\d{1,2}[\-/][a-zA-Z]{3,9}[\-/]\d{2,4}|\d{4}[\-/]\d{1,2}[\-/]\d{1,2})', text_clean, re.IGNORECASE)
                    if m:
                        s_raw = m.group(1).strip()
                        e_raw = m.group(2).strip()
                        year_match = re.search(r'\b(20\d{2})\b', text_clean)
                        year = year_match.group(1) if year_match else str(datetime.datetime.now().year)
                        if not re.search(r'\b20\d{2}\b', s_raw):
                            s_raw = f"{s_raw} {year}"
                        if not re.search(r'\b20\d{2}\b', e_raw):
                            e_raw = f"{e_raw} {year}"
                        detected_validity_start = _clean_date(s_raw)
                        detected_validity_end = _clean_date(e_raw)

                # Contract Number check
                contract_match = re.search(r'(?:contract|service\s*contract|sc)\s*(?:no|number|#)?[:\s]*([A-Za-z0-9\-_/]{4,})', row_text, re.IGNORECASE)
                if contract_match and not detected_contract:
                    candidate_c = contract_match.group(1).strip()
                    if candidate_c.lower() not in ("harge", "rate", "ofr", "summary", "standard", "number", "notes"):
                        detected_contract = candidate_c

            # Find Header Row
            header_row_idx = None
            col_map: Dict[str, int] = {}
            fallback_col_map: Dict[str, int] = {}
            container_cols: List[Tuple[int, str]] = []
            is_lcl_format = False
            lcl_rate_col: Optional[int] = None

            for r in range(1, min(max_row + 1, 25)):
                row_headers = [str(ws.cell(r, c).value or "").strip() for c in range(1, max_col + 1)]
                
                candidate_map: Dict[str, int] = {}
                candidate_fallbacks: Dict[str, int] = {}
                candidate_containers: List[Tuple[int, str]] = []
                candidate_lcl_col: Optional[int] = None
                origin_found = False
                dest_found = False
                container_found = False
                lcl_found = False

                for c_idx, h in enumerate(row_headers):
                    if not h:
                        continue
                    c_1based = c_idx + 1
                    h_norm = _normalize(h)

                    # Check validity start & end FIRST (prevents "valid from" matching origin "from" and "valid to" matching dest "to")
                    if _match_synonym(h, VALIDITY_START_SYNONYMS) and "validity_start" not in candidate_map:
                        candidate_map["validity_start"] = c_1based
                    elif _match_synonym(h, VALIDITY_END_SYNONYMS) and "validity_end" not in candidate_map:
                        candidate_map["validity_end"] = c_1based

                    # Check origin & load port
                    elif _match_synonym(h, ORIGIN_SYNONYMS):
                        if "load port" in h_norm or "pol" in h_norm:
                            candidate_map["origin"] = c_1based
                            origin_found = True
                        elif "origin" not in candidate_map:
                            candidate_map["origin"] = c_1based
                            origin_found = True
                        else:
                            candidate_fallbacks["origin_alt"] = c_1based
                    elif _match_synonym(h, ORIGIN_CODE_SYNONYMS) and "origin_code" not in candidate_map:
                        candidate_map["origin_code"] = c_1based

                    # Check destination & discharge port
                    elif _match_synonym(h, DESTINATION_SYNONYMS):
                        if "discharge port" in h_norm or "pod" in h_norm:
                            candidate_map["destination"] = c_1based
                            dest_found = True
                        elif "destination" not in candidate_map:
                            candidate_map["destination"] = c_1based
                            dest_found = True
                        else:
                            candidate_fallbacks["dest_alt"] = c_1based
                    elif _match_synonym(h, DEST_CODE_SYNONYMS) and "dest_code" not in candidate_map:
                        candidate_map["dest_code"] = c_1based

                    elif _match_synonym(h, CARRIER_SYNONYMS) and "carrier" not in candidate_map:
                        candidate_map["carrier"] = c_1based
                    elif _match_synonym(h, CURRENCY_SYNONYMS) and "currency" not in candidate_map:
                        candidate_map["currency"] = c_1based
                    elif _match_synonym(h, SERVICE_SYNONYMS) and "service" not in candidate_map:
                        candidate_map["service"] = c_1based
                    elif _match_synonym(h, CONTRACT_SYNONYMS) and "contract" not in candidate_map:
                        candidate_map["contract"] = c_1based
                    elif _match_synonym(h, COMMODITY_SYNONYMS) and "commodity" not in candidate_map:
                        candidate_map["commodity"] = c_1based
                    elif _match_synonym(h, TRANSIT_TIME_SYNONYMS) and "transit_time" not in candidate_map:
                        candidate_map["transit_time"] = c_1based

                    ct = _match_container_type(h)
                    if ct:
                        candidate_containers.append((c_1based, ct))
                        container_found = True

                    if _match_lcl_rate(h) and not candidate_lcl_col:
                        candidate_lcl_col = c_1based
                        lcl_found = True

                if (origin_found or dest_found) and not container_found and r + 1 <= max_row:
                    next_row_headers = [str(ws.cell(r + 1, c).value or "").strip() for c in range(1, max_col + 1)]
                    for c_idx, h in enumerate(next_row_headers):
                        ct = _match_container_type(h)
                        if ct:
                            candidate_containers.append((c_idx + 1, ct))
                            container_found = True

                if (origin_found or dest_found) and (container_found or lcl_found):
                    header_row_idx = r
                    col_map = candidate_map
                    fallback_col_map = candidate_fallbacks
                    container_cols = candidate_containers
                    is_lcl_format = lcl_found and not container_found
                    lcl_rate_col = candidate_lcl_col
                    if container_found and any(_match_container_type(str(ws.cell(r + 1, c).value or "")) for c in range(1, max_col + 1)):
                        header_row_idx = r + 1
                    break

            if header_row_idx is None:
                print(f"[GenericExcel] No header row found in sheet '{sheet_name}', skipping.")
                continue

            print(f"[GenericExcel] Sheet '{sheet_name}': header at row {header_row_idx}, col_map={col_map}, containers={container_cols}")

            # Extract Rate Rows
            data_start = header_row_idx + 1
            last_origin_val = ""
            last_origin_code = ""
            last_dest_val = ""
            last_dest_code = ""

            for r in range(data_start, max_row + 1):
                origin_val = str(ws.cell(r, col_map["origin"]).value or "").strip() if "origin" in col_map else ""
                if not origin_val and "origin_alt" in fallback_col_map:
                    origin_val = str(ws.cell(r, fallback_col_map["origin_alt"]).value or "").strip()

                origin_code = str(ws.cell(r, col_map["origin_code"]).value or "").strip() if "origin_code" in col_map else ""

                dest_val = str(ws.cell(r, col_map["destination"]).value or "").strip() if "destination" in col_map else ""
                if not dest_val and "dest_alt" in fallback_col_map:
                    dest_val = str(ws.cell(r, fallback_col_map["dest_alt"]).value or "").strip()

                dest_code = str(ws.cell(r, col_map["dest_code"]).value or "").strip() if "dest_code" in col_map else ""

                if origin_val: last_origin_val = origin_val
                elif not origin_code: origin_val = last_origin_val

                if dest_val: last_dest_val = dest_val
                elif not dest_code: dest_val = last_dest_val

                if not origin_val and not origin_code and not dest_val and not dest_code:
                    continue

                if _is_junk_rate_row(origin_val, dest_val, 0):
                    continue

                carrier = carrier_from_filename or detected_carrier
                if "carrier" in col_map:
                    c_val = str(ws.cell(r, col_map["carrier"]).value or "").strip()
                    if c_val: carrier = c_val

                currency = detected_currency
                if "currency" in col_map:
                    c_val = str(ws.cell(r, col_map["currency"]).value or "").strip().upper()
                    if c_val and len(c_val) == 3: currency = c_val

                service = str(ws.cell(r, col_map["service"]).value or "").strip() if "service" in col_map else ""
                
                v_start = detected_validity_start
                v_end = detected_validity_end
                if "validity_start" in col_map:
                    vs = ws.cell(r, col_map["validity_start"]).value
                    if vs: v_start = _clean_date(vs)
                if "validity_end" in col_map:
                    ve = ws.cell(r, col_map["validity_end"]).value
                    if ve: v_end = _clean_date(ve)

                contract = detected_contract
                if "contract" in col_map:
                    c_val = str(ws.cell(r, col_map["contract"]).value or "").strip()
                    if c_val and len(c_val) >= 4 and c_val.lower() not in ("harge", "rate", "ofr", "summary"):
                        contract = c_val

                commodity = "FAK"
                if "commodity" in col_map:
                    c_val = str(ws.cell(r, col_map["commodity"]).value or "").strip()
                    if c_val: commodity = c_val

                tt_val = str(ws.cell(r, col_map["transit_time"]).value or "").strip() if "transit_time" in col_map else ""

                # Semicolon and Newline multi-port splitting
                raw_d_splits = re.split(r'[\r\n;]+', dest_val) if dest_val else [""]
                dest_list = [d.strip() for d in raw_d_splits if d.strip()]
                if not dest_list:
                    dest_list = [""]

                raw_o_splits = re.split(r'[\r\n;]+', origin_val) if origin_val else [""]
                origin_list = [o.strip() for o in raw_o_splits if o.strip()]
                if not origin_list:
                    origin_list = [""]

                for orig in origin_list:
                    for dest in dest_list:
                        if _is_junk_rate_row(orig, dest, 0):
                            continue

                        route_surcharges = self._match_surcharges(surcharges_cache, orig, dest)

                        for col_idx, container_type in container_cols:
                            amt = _extract_amount(ws.cell(r, col_idx).value)
                            if amt is not None and amt > 0:
                                row_charges = [
                                    ChargeItem(charge_code="BAS", charge_name="Base Ocean Freight", amount=amt, currency=currency, basis="per equipment")
                                ]
                                extra_surcharges = route_surcharges.get(container_type, [])
                                row_charges.extend(extra_surcharges)

                                rate_row = RateRow(
                                    row_index=row_counter,
                                    carrier_scac=carrier.upper() if carrier else "UNKN",
                                    origin_raw=orig or origin_code,
                                    origin_locode=origin_code or orig,
                                    destination_raw=dest or dest_code,
                                    destination_locode=dest_code or dest,
                                    service_type=service,
                                    cargo_type=commodity,
                                    load_type=container_type,
                                    commodity=commodity,
                                    ofr_amount=amt,
                                    ofr_currency=currency,
                                    charges=row_charges,
                                    validity_start=v_start,
                                    validity_end=v_end,
                                    contract_number=contract,
                                    remarks=tt_val,
                                    internal_remarks=f"Transit Time: {tt_val}" if tt_val else ""
                                )
                                all_rates.append(rate_row)
                                row_counter += 1

        carrier_code = carrier_from_filename or detected_carrier or "UNKN"
        summary = JobSummary(
            total_rows=len(all_rates),
            carriers_found=[carrier_code] if carrier_code else ["UNKN"]
        )

        print(f"[GenericExcel] Extracted {len(all_rates)} rate rows from {file_path.name}")

        # Fallback contract and validity from rows if not detected in top rows
        if not detected_contract and all_rates and all_rates[0].contract_number:
            detected_contract = all_rates[0].contract_number
        if not detected_validity_start and all_rates and all_rates[0].validity_start:
            detected_validity_start = all_rates[0].validity_start
            detected_validity_end = all_rates[0].validity_end

        # Check sheet title for contract reference like "TO_USA PORTS"
        if not detected_contract and len(wb.sheetnames) > 0:
            first_cell = str(wb[wb.sheetnames[0]].cell(1, 1).value or "").strip()
            title_m = re.match(r'^([A-Za-z0-9\-_/\s]{3,30}?)(?:\s*\(|$|\s*\-)', first_cell)
            if title_m:
                candidate = title_m.group(1).strip()
                if is_valid_contract_number(candidate):
                    detected_contract = candidate

        summary.contract_number = detected_contract
        summary.validity_start = detected_validity_start
        summary.validity_end = detected_validity_end

        return CanonicalRateSheet(
            job_id=job_id,
            file_name=file_path.name,
            carrier_code=carrier_code,
            contract_number=detected_contract,
            validity_start=detected_validity_start,
            validity_end=detected_validity_end,
            rates=all_rates,
            summary=summary,
        )

    def _extract_surcharge_sheet_data(self, wb, surcharge_sheet_names: List[str]) -> Dict[str, Dict[str, List[ChargeItem]]]:
        surcharges_by_key: Dict[str, Dict[str, List[ChargeItem]]] = {}

        for sname in surcharge_sheet_names:
            ws = wb[sname]
            if ws.max_row is None or ws.max_row < 2:
                continue

            max_col = min(ws.max_column or 1, 35)
            max_row = min(ws.max_row or 1, 2000)

            hdr_row = None
            dest_col = None
            orig_col = None
            desc_col = None
            code_col = None
            curr_col = None
            containers: List[Tuple[int, str]] = []

            for r in range(1, min(max_row + 1, 10)):
                row_vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, max_col + 1)]
                for c_idx, h in enumerate(row_vals):
                    c_1 = c_idx + 1
                    h_norm = _normalize(h)
                    if "discharge" in h_norm or "destination" in h_norm or "dest" in h_norm or "pod" in h_norm:
                        dest_col = c_1
                    elif "load" in h_norm or "origin" in h_norm or "pol" in h_norm:
                        orig_col = c_1
                    elif "description" in h_norm or "surcharge name" in h_norm:
                        desc_col = c_1
                    elif "code" in h_norm or "surcharge code" in h_norm:
                        code_col = c_1
                    elif "curr" in h_norm or "currency" in h_norm:
                        curr_col = c_1
                    
                    ct = _match_container_type(h)
                    if ct:
                        containers.append((c_1, ct))

                if (dest_col or orig_col) and (desc_col or code_col) and containers:
                    hdr_row = r
                    break

            if hdr_row is None:
                continue

            last_dest = ""
            for r in range(hdr_row + 1, max_row + 1):
                d_val = str(ws.cell(r, dest_col).value or "").strip() if dest_col else ""
                if d_val: last_dest = d_val
                else: d_val = last_dest

                desc = str(ws.cell(r, desc_col).value or "").strip() if desc_col else ""
                code = str(ws.cell(r, code_col).value or "").strip() if code_col else ""
                curr = str(ws.cell(r, curr_col).value or "USD").strip().upper() if curr_col else "USD"

                if not desc and not code:
                    continue
                if desc.lower() in ("container charges", "rate offer per container", "subject to", "b/l charges"):
                    continue

                charge_code = code or desc[:4].upper().replace(" ", "")
                charge_name = desc or code

                for c_idx, ct in containers:
                    amt = _extract_amount(ws.cell(r, c_idx).value)
                    if amt is not None and amt > 0:
                        chg_item = ChargeItem(
                            charge_code=charge_code,
                            charge_name=charge_name,
                            amount=amt,
                            currency=curr,
                            basis="per equipment",
                            category="Origin" if "OTHC" in charge_code or "EXPORT" in charge_name.upper() else "Destination" if "DTHC" in charge_code or "DISCHARGE" in charge_name.upper() else "Freight"
                        )
                        d_clean = _normalize(d_val)
                        surcharges_by_key.setdefault(d_clean, {}).setdefault(ct, []).append(chg_item)

        return surcharges_by_key

    def _match_surcharges(self, cache: Dict[str, Dict[str, List[ChargeItem]]], origin: str, dest: str) -> Dict[str, List[ChargeItem]]:
        d_clean = _normalize(dest)
        if d_clean in cache:
            return cache[d_clean]
        
        for k, v in cache.items():
            if k in d_clean or d_clean in k:
                return v
            tokens = [t for t in re.split(r'\W+', d_clean) if len(t) > 2]
            if any(t in k for t in tokens):
                return v

        return {}

    def _detect_carrier_from_filename(self, fn: str) -> str:
        fn_l = fn.lower()
        if "anl" in fn_l or "autnb" in fn_l: return "ANNU"
        if "oocl" in fn_l: return "OOLU"
        if "maeu" in fn_l or "maersk" in fn_l or "o3e" in fn_l or "o3w" in fn_l: return "MAEU"
        if "msc" in fn_l: return "MSCU"
        if "one" in fn_l: return "ONEY"
        if "cosco" in fn_l: return "COSU"
        if "hapag" in fn_l or "hlcu" in fn_l: return "HLCU"
        return ""

    def _detect_validity_from_filename(self, fn: str) -> Optional[Tuple[str, str]]:
        year_m = re.search(r'\b(20\d{2})\b', fn)
        year = year_m.group(1) if year_m else "2026"
        if "q3" in fn.lower():
            return f"{year}-07-01", f"{year}-09-30"
        if "aug" in fn.lower():
            return f"{year}-08-01", f"{year}-08-31"
        return None

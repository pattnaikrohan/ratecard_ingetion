import openpyxl
import datetime
import re
from pathlib import Path
from typing import List, Optional, Tuple, Dict
from app.services.parsers.base_parser import BaseParser
from app.models.canonical import CanonicalRateSheet, RateRow, ChargeItem, JobSummary

class MaerskPlugin(BaseParser):
    def can_parse(self, file_path: Path, filename: str) -> bool:
        fn = filename.lower()
        if any(k in fn for k in ["maeu", "maersk", "o3e", "o3w", "coastal", "299163347", "299952465", "297111717"]):
            return True
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            sheet_names_lower = [s.lower() for s in wb.sheetnames]
            if any(s in ["ocean rates", "afls quote", "bas+surcharges"] for s in sheet_names_lower):
                wb.close()
                return True
            for sn in wb.sheetnames[:4]:
                ws = wb[sn]
                for r in range(1, 15):
                    cell_a = str(ws.cell(r, 1).value or "").lower()
                    cell_b = str(ws.cell(r, 2).value or "").lower()
                    if "maersk" in cell_a or "maersk" in cell_b or "afls" in cell_a or "coastal bridge" in cell_b:
                        wb.close()
                        return True
            wb.close()
        except Exception:
            pass
        return False

    def parse(self, file_path: Path, job_id: str) -> CanonicalRateSheet:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        contract_no = ""
        valid_start = ""
        valid_end = ""

        # Check for Title Page
        title_sheet = None
        for sn in wb.sheetnames:
            if "title" in sn.lower() or "cover" in sn.lower():
                title_sheet = wb[sn]
                break
        
        if title_sheet:
            for r in range(1, min(title_sheet.max_row + 1, 20)):
                k = str(title_sheet.cell(r, 1).value or "").strip().lower()
                v = title_sheet.cell(r, 2).value
                if not v:
                    continue
                v_str = str(v).strip()
                if "contract" in k and ("no" in k or "number" in k):
                    contract_no = v_str
                elif "effective" in k:
                    valid_start = self._clean_date(v)
                elif "expiry" in k or "expire" in k:
                    valid_end = self._clean_date(v)

        # Detect primary rate sheet
        sheet_name = None
        for preferred in ["Ocean Rates", "AFLS Quote", "BAS+Surcharges"]:
            for sn in wb.sheetnames:
                if sn.strip().lower() == preferred.lower():
                    sheet_name = sn
                    break
            if sheet_name:
                break

        if not sheet_name:
            for sn in wb.sheetnames:
                if "title" in sn.lower() or "abbreviation" in sn.lower() or "disclaimer" in sn.lower():
                    continue
                ws_check = wb[sn]
                if ws_check.max_row and ws_check.max_row > 10:
                    sheet_name = sn
                    break

        if not sheet_name:
            sheet_name = wb.sheetnames[0]
            
        ws = wb[sheet_name]
        print(f"[Maersk] Parsing sheet: {sheet_name} (rows={ws.max_row})")

        # Scan top rows of rate sheet for metadata if not found in Title Page
        for r in range(1, min(ws.max_row + 1, 10)):
            cell_a = str(ws.cell(r, 1).value or "").strip()
            cell_b = ws.cell(r, 2).value
            if not cell_b:
                continue
            cell_b_str = str(cell_b).strip()
            if "Contract Number" in cell_a and not contract_no:
                contract_no = cell_b_str
            elif ("Last Acceptance Date" in cell_a or "Effective Date" in cell_a) and not valid_start:
                valid_start = self._clean_date(cell_b)
            elif ("Expiry Date" in cell_a or "Valid To" in cell_a) and not valid_end:
                valid_end = self._clean_date(cell_b)

        # Find header row
        header_row_idx = 7
        for r in range(1, min(ws.max_row + 1, 20)):
            row_vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, min(ws.max_column + 1, 30))]
            if any("receipt" in v or "port of loading" in v or "pol" in v for v in row_vals) and any("charge" in v for v in row_vals):
                header_row_idx = r
                break

        col_layout = self._detect_columns(ws, header_row_idx)
        print(f"[Maersk] Header at row {header_row_idx}, col_layout={col_layout}")

        # ── Phase 1: Collect Surcharges by Route and Container Type ──
        surcharges_map: Dict[str, Dict[str, List[ChargeItem]]] = {} # route_key -> container_type -> [ChargeItem]
        
        for r in range(header_row_idx + 1, min(ws.max_row + 1, 10000)):
            chg_type = str(ws.cell(r, col_layout["charge"]).value or "").strip().upper()
            if not chg_type or chg_type == "BAS":
                continue

            from_loc = str(ws.cell(r, col_layout["origin"]).value or "").strip()
            to_loc = str(ws.cell(r, col_layout["destination"]).value or "").strip()
            route_key = f"{from_loc.upper()}-->{to_loc.upper()}"

            if route_key not in surcharges_map:
                surcharges_map[route_key] = {}

            default_curr = "USD"
            if col_layout.get("currency"):
                default_curr = str(ws.cell(r, col_layout["currency"]).value or "USD").strip().upper()

            basis = "per equipment"
            if col_layout.get("rate_basis"):
                basis_val = str(ws.cell(r, col_layout["rate_basis"]).value or "").strip()
                if "DOCUMENT" in basis_val.upper() or "PER_DOCUMENT" in basis_val.upper():
                    basis = "per document"

            for eq_type, col_idx in col_layout.get("containers", []):
                cell_val = ws.cell(r, col_idx).value
                if cell_val is None:
                    continue
                amt, curr = self._extract_amount_currency(cell_val, default_curr)
                if amt is not None and amt > 0:
                    chg_item = ChargeItem(
                        charge_code=chg_type,
                        charge_name=self._get_charge_description(chg_type),
                        amount=amt,
                        currency=curr,
                        basis=basis,
                        category="Origin" if chg_type.startswith(("O", "OPA", "OPE")) else "Destination" if chg_type.startswith(("D", "DPA", "DPE")) else "Freight"
                    )
                    surcharges_map[route_key].setdefault(eq_type, []).append(chg_item)

        # ── Phase 2: Extract BAS Rates and Attach Surcharges ──
        rates: List[RateRow] = []
        row_counter = 1

        for r in range(header_row_idx + 1, min(ws.max_row + 1, 10000)):
            chg_type = str(ws.cell(r, col_layout["charge"]).value or "").strip().upper()
            if chg_type != "BAS":
                continue

            from_loc = str(ws.cell(r, col_layout["origin"]).value or "").strip()
            to_loc = str(ws.cell(r, col_layout["destination"]).value or "").strip()
            route_key = f"{from_loc.upper()}-->{to_loc.upper()}"
            
            eff_dt = valid_start
            exp_dt = valid_end
            if col_layout.get("effective"):
                eff_dt = self._clean_date(ws.cell(r, col_layout["effective"]).value) or valid_start
            if col_layout.get("expiry"):
                exp_dt = self._clean_date(ws.cell(r, col_layout["expiry"]).value) or valid_end

            svc_mode = "CY/CY"
            if col_layout.get("service_mode"):
                svc_mode = str(ws.cell(r, col_layout["service_mode"]).value or "CY/CY").strip()
            
            comm = "FAK"
            if col_layout.get("commodity"):
                comm = str(ws.cell(r, col_layout["commodity"]).value or "FAK").strip()

            tt_str = ""
            if col_layout.get("transit_time"):
                tt_val = str(ws.cell(r, col_layout["transit_time"]).value or "").strip()
                if tt_val and tt_val.lower() != "none":
                    tt_str = tt_val

            default_curr = "USD"
            if col_layout.get("currency"):
                default_curr = str(ws.cell(r, col_layout["currency"]).value or "USD").strip().upper()

            for eq_type, col_idx in col_layout.get("containers", []):
                cell_val = ws.cell(r, col_idx).value
                if cell_val is None:
                    continue
                
                amt, curr = self._extract_amount_currency(cell_val, default_curr)
                if amt is None or amt <= 0:
                    continue

                # Collect charges for this route & container type
                row_charges = [
                    ChargeItem(charge_code="BAS", charge_name="Base Ocean Freight", amount=amt, currency=curr, basis="per equipment")
                ]
                extra_surcharges = surcharges_map.get(route_key, {}).get(eq_type, [])
                row_charges.extend(extra_surcharges)

                r_row = RateRow(
                    row_index=row_counter,
                    carrier_scac="MAEU",
                    origin_raw=from_loc,
                    origin_locode=from_loc,
                    destination_raw=to_loc,
                    destination_locode=to_loc,
                    service_type=svc_mode,
                    cargo_type="FAK",
                    load_type=eq_type,
                    commodity=comm,
                    ofr_amount=amt,
                    ofr_currency=curr,
                    charges=row_charges,
                    validity_start=eff_dt,
                    validity_end=exp_dt,
                    contract_number=contract_no,
                    remarks=tt_str,
                    internal_remarks=f"Transit Time: {tt_str}" if tt_str else ""
                )
                rates.append(r_row)
                row_counter += 1

        print(f"[Maersk] Extracted {len(rates)} BAS rate rows with attached surcharges from {sheet_name}")
        summary = JobSummary(total_rows=len(rates), carriers_found=["MAEU"])
        return CanonicalRateSheet(
            job_id=job_id,
            file_name=file_path.name,
            carrier_code="MAEU",
            contract_number=contract_no,
            validity_start=valid_start,
            validity_end=valid_end,
            rates=rates,
            summary=summary
        )

    def _get_charge_description(self, code: str) -> str:
        descriptions = {
            "PAI": "Port Additional Import",
            "PAE": "Port Additional Export",
            "VPI": "Vessel Peak Season Import",
            "EBS": "Emergency Bunker Surcharge",
            "DDF": "Destination Documentation Fee",
            "ODF": "Origin Documentation Fee",
            "CFD": "Container Cleaning Fee Destination",
            "DTHC": "Destination Terminal Handling Charge",
            "OTHC": "Origin Terminal Handling Charge",
            "BAF": "Bunker Adjustment Factor",
            "LSS": "Low Sulfur Surcharge",
            "EFF": "Environmental Fuel Fee",
            "OHC": "Origin Handling Charge",
            "DPA": "Destination Port Additional",
            "OPA": "Origin Port Additional"
        }
        return descriptions.get(code, f"Surcharge ({code})")

    def _detect_columns(self, ws, header_row: int) -> dict:
        layout = {
            "origin": 1,
            "destination": 4,
            "charge": 10,
            "containers": [],
        }
        
        headers = {}
        for c in range(1, min((ws.max_column or 1) + 1, 30)):
            val = str(ws.cell(header_row, c).value or "").strip().lower()
            if val:
                headers[c] = val

        for col_idx, h in headers.items():
            h_clean = re.sub(r'\s+', ' ', h).strip()
            
            if h_clean in ("receipt", "por", "place of receipt", "origin") or h_clean.startswith("receipt"):
                layout["origin"] = col_idx
            elif h_clean in ("delivery", "del", "place of delivery", "destination", "pod") or h_clean.startswith("delivery"):
                layout["destination"] = col_idx
            elif h_clean in ("pol", "port of loading"):
                layout.setdefault("pol", col_idx)
            elif h_clean in ("pod", "port of discharge"):
                layout.setdefault("pod_col", col_idx)
            elif "effective" in h_clean:
                layout["effective"] = col_idx
            elif "expiry" in h_clean or "expire" in h_clean:
                layout["expiry"] = col_idx
            elif "transit" in h_clean or "tt" in h_clean:
                layout["transit_time"] = col_idx
            elif "rate basis" in h_clean or "basis" in h_clean:
                layout["rate_basis"] = col_idx
            elif "service" in h_clean or "mode" in h_clean or "sm" in h_clean:
                layout["service_mode"] = col_idx
            elif "commodity" in h_clean:
                layout["commodity"] = col_idx
            elif h_clean in ("charge", "chg", "charge code"):
                layout["charge"] = col_idx
            elif h_clean in ("currency", "curr"):
                layout["currency"] = col_idx
            elif "20" in h_clean and ("dry" in h_clean or "gp" in h_clean or "dc" in h_clean or "st" in h_clean):
                layout["containers"].append(("20GP", col_idx))
            elif "40h" in h_clean or "40hd" in h_clean:
                layout["containers"].append(("40HC", col_idx))
            elif "40" in h_clean and ("dry" in h_clean or "gp" in h_clean or "dc" in h_clean or "st" in h_clean):
                layout["containers"].append(("40GP", col_idx))
            elif "45" in h_clean:
                layout["containers"].append(("45HC", col_idx))

        if not layout["containers"]:
            layout["containers"] = [("20GP", 13), ("40GP", 14), ("40HC", 15)]

        return layout

    def _extract_amount_currency(self, val, default_curr: str = "USD") -> Tuple[Optional[float], str]:
        if val is None:
            return None, default_curr
        s = str(val).strip()
        if not s or s.lower() in ("none", "null", "-", "n/a"):
            return None, default_curr
        try:
            return float(s), default_curr
        except ValueError:
            pass
        match = re.match(r'^([\d.,]+)\s*([A-Z]{3})$', s.strip())
        if match:
            try:
                amt = float(match.group(1).replace(',', ''))
                return amt, match.group(2)
            except ValueError:
                pass
        match = re.match(r'^([A-Z]{3})\s*([\d.,]+)$', s.strip())
        if match:
            try:
                amt = float(match.group(2).replace(',', ''))
                return amt, match.group(1)
            except ValueError:
                pass
        return None, default_curr

    def _clean_date(self, date_val) -> str:
        if not date_val:
            return ""
        if isinstance(date_val, (datetime.datetime, datetime.date)):
            return date_val.strftime("%Y-%m-%d")
        s = str(date_val).strip()
        s = re.sub(r'\s+\d{2}:\d{2}:\d{2}.*$', '', s)
        for fmt in ["%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"]:
            try:
                dt = datetime.datetime.strptime(s, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                pass
        return s

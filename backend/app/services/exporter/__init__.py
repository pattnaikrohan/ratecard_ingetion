import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path
from typing import List
from app.core.config import TEMPLATE_XLSM, PROCESSED_DIR
from app.models.canonical import CanonicalRateSheet, RateRow

class FreightifyExporter:
    def __init__(self, template_path: Path = TEMPLATE_XLSM):
        self.template_path = template_path
        self.clean_template_path = PROCESSED_DIR / "clean_template.xlsm"

    def _ensure_clean_template(self) -> Path:
        """Create a lightweight macro-enabled template containing ONLY the 'Template' sheet."""
        if self.clean_template_path.exists():
            return self.clean_template_path

        PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        print(f"[Exporter] Generating clean export template at {self.clean_template_path}...")
        try:
            wb = openpyxl.load_workbook(self.template_path, keep_vba=True)
            for s in list(wb.sheetnames):
                if s != "Template":
                    del wb[s]
            ws = wb["Template"]
            # Prune all sample data rows (keep only Row 1 header)
            ws._cells = {(1, c): cell for (r, c), cell in ws._cells.items() if r == 1}
            wb.save(self.clean_template_path)
            print(f"[Exporter] Clean template created successfully ({self.clean_template_path.stat().st_size / 1024:.1f} KB)")
            return self.clean_template_path
        except Exception as e:
            print(f"[Exporter] Failed to create clean template: {e}, falling back to original")
            return self.template_path

    def export(self, sheet: CanonicalRateSheet, output_filename: str, export_policy: str = "PARTIAL") -> Path:
        print(f"Exporting Canonical Rate Sheet ({len(sheet.rates)} rows) to Freightify Workbook {output_filename}...")
        
        # Load lightweight clean template
        template_file = self._ensure_clean_template()
        wb = openpyxl.load_workbook(template_file, keep_vba=True)
        ws = wb["Template"]

        # Ensure worksheet only contains row 1 headers (no ghost rows)
        ws._cells = {(1, c): cell for (r, c), cell in ws._cells.items() if r == 1}

        # Styling definitions for visual highlighting in exported Excel
        warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft gold/yellow
        error_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft rose/red

        # Filter rows based on export policy
        rows_to_export: List[RateRow] = []
        for r in sheet.rates:
            if export_policy == "STRICT":
                if r.validation_status in ["VALID", "INFO", "WARNING"]:
                    rows_to_export.append(r)
            elif export_policy == "PARTIAL":
                if r.validation_status != "CRITICAL":
                    rows_to_export.append(r)
            else:  # WARNING_PERMISSIVE
                rows_to_export.append(r)

        for idx, rate in enumerate(rows_to_export, start=2):
            c_scac = ws.cell(row=idx, column=1, value=rate.carrier_scac)
            c_orig = ws.cell(row=idx, column=2, value=rate.origin_locode)
            ws.cell(row=idx, column=3, value=rate.origin_locode)
            c_dest = ws.cell(row=idx, column=5, value=rate.destination_locode)
            ws.cell(row=idx, column=7, value=rate.service_type or "")
            ws.cell(row=idx, column=8, value=rate.cargo_type or "FAK")
            ws.cell(row=idx, column=10, value=rate.commodity or "")
            ws.cell(row=idx, column=12, value=rate.inclusions or "")
            ws.cell(row=idx, column=13, value=rate.subject_to or "")
            
            # Remarks column notes validation status and reasons
            remarks_text = rate.remarks or ""
            if rate.validation_status == "WARNING" and rate.validation_messages:
                warning_notes = "; ".join(rate.validation_messages)
                remarks_text = f"[WARNING: {warning_notes}] {remarks_text}".strip()
            elif rate.validation_status == "ERROR" and rate.validation_messages:
                error_notes = "; ".join(rate.validation_messages)
                remarks_text = f"[ERROR: {error_notes}] {remarks_text}".strip()
            ws.cell(row=idx, column=14, value=remarks_text)

            c_load = ws.cell(row=idx, column=16, value=rate.load_type)
            c_vstart = ws.cell(row=idx, column=17, value=rate.validity_start or sheet.validity_start or "")
            c_vend = ws.cell(row=idx, column=18, value=rate.validity_end or sheet.validity_end or "")
            ws.cell(row=idx, column=21, value=rate.contract_number or sheet.contract_number or "")
            
            # Base Ocean Freight (OFR)
            c_ofr = ws.cell(row=idx, column=22, value=rate.ofr_amount)
            ws.cell(row=idx, column=23, value="per equipment")
            ws.cell(row=idx, column=24, value=rate.ofr_currency or "USD")

            # Apply soft pastel fills based on validation status
            key_cells = [c_scac, c_orig, c_dest, c_load, c_vstart, c_vend, c_ofr]
            if rate.validation_status == "WARNING":
                for kc in key_cells:
                    kc.fill = warning_fill
            elif rate.validation_status == "ERROR":
                for kc in key_cells:
                    kc.fill = error_fill

            # Standard Freightify Surcharge Column Mapping
            for chg in rate.charges:
                code = (chg.charge_code or "").upper().strip()
                if code in ["CGS (DESTINATION)", "CGS_DEST", "CGS"]:
                    ws.cell(row=idx, column=27, value=chg.amount)
                    ws.cell(row=idx, column=28, value=chg.basis or "per equipment")
                    ws.cell(row=idx, column=29, value=chg.currency or rate.ofr_currency or "USD")
                elif code in ["CGS (ORIGIN)", "CGS_ORIG"]:
                    ws.cell(row=idx, column=32, value=chg.amount)
                    ws.cell(row=idx, column=33, value=chg.basis or "per equipment")
                    ws.cell(row=idx, column=34, value=chg.currency or rate.ofr_currency or "USD")
                elif code in ["DOC", "DOC (DESTINATION)", "DOC_DEST"]:
                    ws.cell(row=idx, column=37, value=chg.amount)
                    ws.cell(row=idx, column=38, value=chg.basis or "per B/L")
                    ws.cell(row=idx, column=39, value=chg.currency or "AUD")
                elif code in ["THC", "THC (DESTINATION)", "DTHC", "THC_DEST"]:
                    ws.cell(row=idx, column=42, value=chg.amount)
                    ws.cell(row=idx, column=43, value=chg.basis or "per equipment")
                    ws.cell(row=idx, column=44, value=chg.currency or "AUD")
                elif code in ["EXP", "EXP (ORIGIN)", "EXP_ORIG"]:
                    ws.cell(row=idx, column=47, value=chg.amount)
                    ws.cell(row=idx, column=48, value=chg.basis or "per equipment")
                    ws.cell(row=idx, column=49, value=chg.currency or "CNY")
                elif code in ["DOC (ORIGIN)", "DOC_ORIG"]:
                    ws.cell(row=idx, column=52, value=chg.amount)
                    ws.cell(row=idx, column=53, value=chg.basis or "per B/L")
                    ws.cell(row=idx, column=54, value=chg.currency or "CNY")
                elif code in ["THC (ORIGIN)", "OTHC", "THC_ORIG"]:
                    ws.cell(row=idx, column=57, value=chg.amount)
                    ws.cell(row=idx, column=58, value=chg.basis or "per equipment")
                    ws.cell(row=idx, column=59, value=chg.currency or "CNY")
                elif code in ["VP1", "VP1 (FREIGHT)", "VP1_FRT"]:
                    ws.cell(row=idx, column=62, value=chg.amount)
                    ws.cell(row=idx, column=63, value=chg.basis or "per equipment")
                    ws.cell(row=idx, column=64, value=chg.currency or "USD")

        output_path = PROCESSED_DIR / output_filename
        wb.save(output_path)
        print(f"Exported {len(rows_to_export)} rows to {output_path}")
        return output_path
